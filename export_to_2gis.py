#!/usr/bin/env python3
"""
Экспорт товаров МойСклад → Excel-шаблон 2ГИС
==============================================
Фильтр: по артикулам (ARTICLES)
Формат: xlsx (шаблон 2ГИС)

Использование:
    pip install requests openpyxl
    python export_to_2gis.py

Результат: файл 2gis_export.xlsx
Загрузи его в: https://account.2gis.com → Товары и услуги → Из файла → Excel
"""

import requests
import sys
import os
import base64
import openpyxl
from openpyxl.styles import Font
from datetime import datetime
from urllib.parse import quote

# Windows-консоль по умолчанию cp1251 и падает на символах ✓/📷 — форсируем UTF-8
try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except (AttributeError, ValueError):
    pass

# ── Загрузка .env ────────────────────────────────────────────────────────────────
def load_dotenv(path: str = ".env") -> None:
    """Простой загрузчик .env: строки вида KEY=VALUE → os.environ (без перезаписи)."""
    if not os.path.exists(path):
        return
    with open(path, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            key, _, value = line.partition("=")
            os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))


load_dotenv()

# ── Настройки ──────────────────────────────────────────────────────────────────
# Секреты берутся из переменных окружения (.env), а НЕ хранятся в коде.
TOKEN        = os.environ.get("MOYSKLAD_TOKEN", "")
GITHUB_TOKEN = os.environ.get("GITHUB_TOKEN", "")

ARTICLES = ["g1007"]
SHOP_URL          = "https://example.com"   # замени на свой сайт
CATEGORY_OVERRIDE = "Столы шпонированные"   # если пусто — берётся из МойСклад

GITHUB_REPO  = "ilyaizhanov-prog/2gis"
GITHUB_PAGES = "https://ilyaizhanov-prog.github.io/2gis"
OUTPUT   = "2gis_export.xlsx"

if not TOKEN:
    sys.exit("❌ Не задан MOYSKLAD_TOKEN. Укажи его в файле .env (см. .env.example).")
if not GITHUB_TOKEN:
    sys.exit("❌ Не задан GITHUB_TOKEN. Укажи его в файле .env (см. .env.example).")

BASE_URL = "https://api.moysklad.ru/api/remap/1.2"
HEADERS  = {
    "Authorization": f"Bearer {TOKEN}",
    "Accept-Encoding": "gzip",
}


# ── Получение данных ───────────────────────────────────────────────────────────

def get_product_by_article(article: str) -> dict | None:
    """Находит базовый товар по артикулу."""
    resp = requests.get(
        f"{BASE_URL}/entity/product",
        headers=HEADERS,
        params={"filter": f"article={article}", "expand": "productFolder", "limit": 1},
        timeout=15,
    )
    resp.raise_for_status()
    rows = resp.json().get("rows", [])
    return rows[0] if rows else None


def get_variants(product_id: str) -> list:
    """Получает все модификации товара."""
    resp = requests.get(
        f"{BASE_URL}/entity/variant",
        headers=HEADERS,
        params={"filter": f"productid={product_id}", "limit": 100},
        timeout=15,
    )
    resp.raise_for_status()
    return resp.json().get("rows", [])


def get_products(articles: list) -> tuple:
    """Возвращает (базовые товары, модификации) по списку артикулов."""
    all_base = []
    all_variants = []

    for article in articles:
        base = get_product_by_article(article)
        if not base:
            print(f"  ✗ Артикул {article}: не найден")
            continue

        pid = base["id"]
        variants = get_variants(pid)

        if variants:
            print(f"  ✓ Артикул {article}: найдено {len(variants)} модификаций")
            all_variants.extend(variants)
            all_base.append(base)
        else:
            print(f"  ✓ Артикул {article}: модификаций нет, берём базовый товар")
            all_base.append(base)

    return all_base, all_variants


def get_sale_price(obj: dict) -> float | None:
    """Цена продажи (тип 'Цена продажи' или первая доступная)."""
    prices = obj.get("salePrices", [])
    # Сначала ищем именно "Цена продажи"
    for price in prices:
        ptype = price.get("priceType", {}).get("name", "")
        if "продаж" in ptype.lower():
            val = price.get("value", 0)
            if val:
                return val / 100
    # Иначе первая ненулевая
    for price in prices:
        val = price.get("value", 0)
        if val:
            return val / 100
    return None


def get_category(product: dict) -> str:
    folder = product.get("productFolder")
    return folder.get("name", "") if folder else ""


def upload_to_github(local_path: str, github_path: str) -> str:
    """Загружает файл в GitHub Pages. Если файл уже есть — пропускает."""
    api_url = f"https://api.github.com/repos/{GITHUB_REPO}/contents/{github_path}"
    gh_headers = {"Authorization": f"Bearer {GITHUB_TOKEN}", "Accept": "application/vnd.github+json"}

    # Публичная ссылка: пробелы/кириллицу нужно URL-кодировать, иначе 2ГИС не загрузит фото
    public_url = f"{GITHUB_PAGES}/{quote(github_path)}"

    # Проверяем — файл уже есть?
    existing = requests.get(api_url, headers=gh_headers, params={"ref": "gh-pages"}, timeout=10)
    if existing.status_code == 200:
        print(f"    ⏭️  Уже есть: {github_path}")
        return public_url

    # Загружаем новый файл
    with open(local_path, "rb") as f:
        content = base64.b64encode(f.read()).decode()

    payload = {"message": f"Фото: {github_path}", "content": content, "branch": "gh-pages"}
    resp = requests.put(api_url, headers=gh_headers, json=payload, timeout=30)
    resp.raise_for_status()
    return public_url


def download_images(product_id: str, name: str, entity: str = "product") -> list:
    """Скачивает фото из МойСклад (product или variant), сохраняет в папку images/."""
    os.makedirs("images", exist_ok=True)
    try:
        resp = requests.get(
            f"{BASE_URL}/entity/{entity}/{product_id}/images",
            headers=HEADERS,
            params={"limit": 5},
            timeout=15,
        )
        resp.raise_for_status()
        rows = resp.json().get("rows", [])
    except Exception:
        return []

    saved = []
    for i, img in enumerate(rows):
        download_href = img.get("meta", {}).get("downloadHref", "")
        if not download_href:
            continue
        try:
            img_resp = requests.get(download_href, headers=HEADERS, timeout=30)
            img_resp.raise_for_status()
            # Определяем расширение по content-type
            ct = img_resp.headers.get("Content-Type", "image/jpeg")
            ext = "jpg" if "jpeg" in ct else ct.split("/")[-1].split(";")[0]
            # Безопасное имя файла
            safe_name = "".join(c for c in name if c.isalnum() or c in " _-")[:40].strip()
            filename = f"images/{safe_name}_{i+1}.{ext}"
            with open(filename, "wb") as f:
                f.write(img_resp.content)

            # Загружаем в GitHub Pages
            try:
                github_path = f"images/{os.path.basename(filename)}"
                public_url = upload_to_github(filename, github_path)
                saved.append(public_url)
                print(f"    📷 Загружено: {public_url}")
            except Exception as e:
                print(f"    ⚠️  Не удалось загрузить в GitHub: {e}")
                saved.append("")
        except Exception:
            continue
    return saved


# ── Запись Excel ───────────────────────────────────────────────────────────────

def build_excel(base_products: list, variants: list) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Лист1"

    headers = [
        "Наименование товара",
        "Цена",
        "Цена от",
        "Цена до",
        "Категория",
        "Ссылка на товар на сайте магазина",
        "Ссылка на картинку",
        "Описание",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    col_widths = [50, 12, 12, 12, 20, 50, 50, 60]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    # Словарь базовых товаров для получения категории и описания
    base_by_id = {p["id"]: p for p in base_products}

    # Если есть модификации — пишем их
    rows_to_write = variants if variants else base_products

    for obj in rows_to_write:
        # Для модификации берём категорию/описание/фото из базового товара
        if variants:
            parent_id = obj.get("product", {}).get("meta", {}).get("href", "").split("/")[-1]
            base = base_by_id.get(parent_id, {})
            category    = CATEGORY_OVERRIDE or get_category(base)
            description = base.get("description", "")
            base_id     = parent_id
            base_name   = base.get("name", "")
        else:
            category    = CATEGORY_OVERRIDE or get_category(obj)
            description = obj.get("description", "")
            base_id     = obj["id"]
            base_name   = obj.get("name", "")

        price  = get_sale_price(obj)
        oid    = obj["id"]
        name   = obj.get("name", "")

        # Сначала ищем фото в модификации, потом в базовом товаре
        images = download_images(oid, name, entity="variant") if variants else []
        if not images:
            images = download_images(base_id, base_name, entity="product")
        picture_url = images[0] if images else ""

        ws.append([
            name,
            price or "",
            "",   # Цена от
            "",   # Цена до
            category,
            "",   # URL товара
            picture_url,
            description,
        ])

    wb.save(OUTPUT)


# ── Точка входа ────────────────────────────────────────────────────────────────

def main():
    print(f"Запрашиваю товары: {ARTICLES}")
    try:
        base_products, variants = get_products(ARTICLES)
    except requests.HTTPError as e:
        print(f"\n❌ Ошибка API МойСклад: {e.response.status_code}")
        print(e.response.text[:300])
        sys.exit(1)

    if not base_products:
        print("\n❌ Товары не найдены. Проверь артикулы в МойСклад.")
        sys.exit(1)

    total = len(variants) if variants else len(base_products)
    print(f"\nВсего строк для экспорта: {total}")
    build_excel(base_products, variants)
    print(f"✅ Файл сохранён: {OUTPUT}")
    print("\nДальше:")
    print("  1. Открой https://account.2gis.com")
    print("  2. Товары и услуги → Из файла → Загрузить файл")
    print(f"  3. Загрузи: {OUTPUT}")
    print("  4. Данные появятся в 2ГИС в течение суток")


if __name__ == "__main__":
    main()
